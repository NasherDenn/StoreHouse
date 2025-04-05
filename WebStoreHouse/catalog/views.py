import json
import time

from django.contrib.auth.decorators import permission_required
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.models import auth
from django.contrib.auth import authenticate
from django.urls import reverse_lazy
from django.views.generic import UpdateView
from django.core.serializers import serialize
from django.http import JsonResponse
from .models import *
from django.views import generic
from .forms import *
# import wadofstuff
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

from openpyxl import Workbook

from django.core.serializers.json import DjangoJSONEncoder

from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, numbers
from openpyxl.worksheet.page import PageMargins
import io
import os
import datetime
from django.conf import settings
import warnings

import logging

from django.db import connection
from django.shortcuts import render


logger = logging.getLogger(__name__)  # Логирование


class UnitList(generic.ListView):
    model = Unit


class MethodList(generic.ListView):
    model = Unit.method


def login(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                auth.login(request, user)
                return redirect('home/')
        else:
            return render(request, 'registration/login.html', {'form': form})
    else:
        form = LoginForm()
    return render(request, 'registration/login.html', {'form': form})


def home(request):
    context_filter_method = MethodNdt.objects.all()
    context_filter_location = Location.objects.all()
    context_filter_status = Status.objects.all()
    # данные об оборудовании из БД
    unit_list = Unit.objects.all()
    unit_form = UnitForm()
    return render(request, 'catalog/home.html', {
        'context_filter_method': context_filter_method,
        'context_filter_location': context_filter_location,
        'context_filter_status': context_filter_status,
        'unit_list': unit_list,
        'form_add': unit_form,
    })


def create(request):
    if request.method == "POST":
        unit = Unit()
        # выбираем из экземпляра значение по выбранному значению по id
        method = MethodNdt.objects.get(id=request.POST.get("method"))
        unit.method = method
        unit.manufacturer = request.POST.get("manufacturer")
        unit.type = request.POST.get("type")
        unit.equipment_name = request.POST.get("name")
        unit.equipment_serial_number = request.POST.get("serial")
        unit.total = request.POST.get("total")
        location = Location.objects.get(id=request.POST.get("location"))
        unit.location = location
        status = Status.objects.get(id=request.POST.get("status"))
        unit.status = status
        unit.notes = request.POST.get("notes")
        unit.save()
        unit.first_id = unit.id
        unit.save()
        # словарь со значениями для формирования истории в БД
        data_write = {'date_write': datetime.datetime.now().strftime('%Y-%m-%d'),
                      'time_write': datetime.datetime.now().time().strftime('%H:%M:%S'),
                      'crud_write': f'добавлено в БД',
                      'from_write': '-',
                      'who_write': request.user.username,
                      'whom_write': '-',
                      'to_write': '-',
                      'method_write': method,
                      'manufacturer_write': request.POST.get("manufacturer"),
                      'type_write': request.POST.get("type"),
                      'name_write': request.POST.get("name"),
                      'serial_number_write': request.POST.get("serial"),
                      'total_write': request.POST.get("total"),
                      'location_write': location,
                      'status_write': status,
                      'notes_write': request.POST.get("notes"),
                      'id_write': unit.id}
        write_history(request, data_write)
        return home(request)


@permission_required('catalog.can_edit_task', raise_exception=True)
def unit_edit(request, id):
    unit = Unit.objects.get(id=id)
    edit_form = EditForm(initial={
        'method': unit.method,
        'manufacturer': unit.manufacturer,
        'type': unit.type,
        'name': unit.equipment_name,
        'serial': unit.equipment_serial_number,
        'total': unit.total,
        'location': unit.location,
        'status': unit.status,
        'notes': unit.notes,
        'id': id,
    })
    return render(request, "catalog/unit_edit.html", {"form_edit": edit_form})


def unit_update(request):
    if request.method == "POST":
        id = request.POST.get("id")
        unit = Unit.objects.get(id=id)
        # выбираем из экземпляра значение по выбранному значению по id
        method = MethodNdt.objects.get(id=request.POST.get("method"))
        unit.method = method
        unit.manufacturer = request.POST.get("manufacturer")
        unit.type = request.POST.get("type")
        unit.equipment_name = request.POST.get("name")
        unit.equipment_serial_number = request.POST.get("serial")
        unit.total = request.POST.get("total")
        location = Location.objects.get(id=request.POST.get("location"))
        unit.location = location
        status = Status.objects.get(id=request.POST.get("status"))
        unit.status = status
        unit.notes = request.POST.get("notes")
        data_write = {}
        # словарь со значениями для формирования истории в БД
        data_write['date_write'] = datetime.datetime.now().strftime('%Y-%m-%d')
        data_write['time_write'] = datetime.datetime.now().time().strftime('%H:%M:%S')
        data_write['crud_write'] = 'внесены изменения'
        data_write['from_write'] = '-'
        data_write['who_write'] = request.user.username
        data_write['whom_write'] = '-'
        data_write['to_write'] = '-'

        if method != Unit.objects.get(id=id).method:
            data_write['method_write'] = f'{Unit.objects.get(id=id).method} &#x2192  {method}'
        else:
            data_write['method_write'] = method

        if request.POST.get("manufacturer") != Unit.objects.get(id=id).manufacturer:
            data_write['manufacturer_write'] = f'{Unit.objects.get(id=id).manufacturer} &#x2192 {request.POST.get("manufacturer")}'
        else:
            data_write['manufacturer_write'] = request.POST.get("manufacturer")

        if request.POST.get("type") != Unit.objects.get(id=id).type:
            data_write['type_write'] = f'{Unit.objects.get(id=id).type} &#x2192 {request.POST.get("type")}'
        else:
            data_write['type_write'] = request.POST.get("type")

        if request.POST.get("name") != Unit.objects.get(id=id).equipment_name:
            data_write['name_write'] = f'{Unit.objects.get(id=id).equipment_name} &#x2192 {request.POST.get("name")}'
        else:
            data_write['name_write'] = request.POST.get("name")

        if request.POST.get("serial") != Unit.objects.get(id=id).equipment_serial_number:
            data_write['serial_number_write'] = f'{Unit.objects.get(id=id).equipment_serial_number} &#x2192 {request.POST.get("serial")}'
        else:
            data_write['serial_number_write'] = request.POST.get("serial")

        if int(request.POST.get("total")) != int(Unit.objects.get(id=id).total):
            data_write['total_write'] = f'{Unit.objects.get(id=id).total} &#x2192 {request.POST.get("total")}'
        else:
            data_write['total_write'] = request.POST.get("total")

        if location != Unit.objects.get(id=id).location:
            data_write['location_write'] = f'{Unit.objects.get(id=id).location} &#x2192 {location}'
        else:
            data_write['location_write'] = location

        if status != Unit.objects.get(id=id).status:
            data_write['status_write'] = f'{Unit.objects.get(id=id).status} &#x2192 {status}'
        else:
            data_write['status_write'] = status

        if request.POST.get("notes") != Unit.objects.get(id=id).notes:
            data_write['notes_write'] = f'{Unit.objects.get(id=id).notes} &#x2192 {request.POST.get("notes")}'
        else:
            data_write['notes_write'] = request.POST.get("notes")

        data_write['id_write'] = Unit.objects.get(id=id).first_id
        unit.save()
        write_history(request, data_write)
        return HttpResponseRedirect("/home/#edit-tab-pane")
        # return HttpResponseRedirect("/home/")


def unit_update_send(request, list_id: list, recip: str):
    '''
    Обновляем данные в БД
    :param list_id: список id из form-send.html, переменной tableData
    :param recip: значение из select form-send.html - получатель оборудования
    '''
    if request.method == "POST":
        for id_index in list_id:
            unit = Unit.objects.get(id=id_index)
            location = Location.objects.get(name=recip)
            unit.location = location
            unit.save()
        return HttpResponse()
    else:
        return HttpResponse("Метод запроса должен быть POST.", status=405)


def check_count_send_equipment(request, list_id: list, table_data: dict):
    '''
    Проверяем какое количество оборудования перемещается.
    1 Если в наличии больше, чем 1 штука (например 6 штук), а отправляется 4, то добавляем дополнительную строку с таким же оборудованием (новый id),
    но с другим количеством и местоположением. А в первой строке уменьшаем значение количества.
    2 Если в месте куда перемещается уже есть точно такое же оборудование (одинаковый метод контроля, производитель, тип, название),
    то увеличиваем количество оборудования в месте перемещения, сливаем вместе примечание, оставляем статус как в месте куда перемещается
    и удаляем оборудование (id), которое перемещаем.
    :param list_id: список id из form-send.html, переменной tableData
    :param table_data: отправляемые данные из таблицы tableData в form-send.html
    '''
    # переменная куда отправляется оборудование
    logging.error(f'Всё что передалось с frontend: {table_data}')
    recip = table_data[-1]['recipient']
    recip_name = table_data[-1]['recipient_name']
    table_data = table_data[:-1]
    logging.error(f'Всё оборудование: {table_data}')
    if request.method == "POST":
        # обходим отправляемое оборудование по id
        for index, index_id in enumerate(list_id):
            logging.error(f'index: {index}')
            logging.error(f'index_id: {index_id}')
            # словарь со значениями для формирования истории в БД
            data_write = {}
            # количество оборудования в БД
            count_unit_db = Unit.objects.get(id=int(index_id)).total
            # количество перемещаемого оборудования
            count_unit_send = table_data[index]['count']
            #       - передаётся одна позиция или несколько:
            #          - передаётся всё оборудование или часть оборудования
            #              - передаётся в новое место где ещё нет никакого оборудования или там есть хоть какое-нибудь оборудование
            #                  - передаваемое оборудование уже присутствует в месте назначения или такого там ещё нет
            #   + одна позиция - всё оборудование - новое место где ещё нет никакого оборудования
            #   + одна позиция - всё оборудование - место где хоть какое-нибудь оборудование - такое оборудование уже присутствует в месте назначения
            #   + одна позиция - всё оборудование - место где хоть какое-нибудь оборудование - такого оборудования там ещё нет
            #   + одна позиция - часть оборудования - новое место где ещё нет никакого оборудования
            #   + одна позиция - часть оборудования - место где хоть какое-нибудь оборудование - такое оборудование уже присутствует в месте назначения
            #   + одна позиция - часть оборудования - место где хоть какое-нибудь оборудование - такого оборудования там ещё нет
            #   + несколько позиций - всё оборудование - новое место где ещё нет никакого оборудования
            #   + несколько позиций - всё оборудование - место где хоть какое-нибудь оборудование - всё такое передаваемое оборудование уже присутствует в месте назначения
            #   + несколько позиций - всё оборудование - место где хоть какое-нибудь оборудование - всего такого оборудования там ещё нет
            #   + несколько позиций - всё оборудование - место где хоть какое-нибудь оборудование - часть такого передаваемого оборудования уже присутствует в месте назначения
            #   + несколько позиций - часть оборудования - новое место где ещё нет никакого оборудования
            #   + несколько позиций - часть оборудования - место где хоть какое-нибудь оборудование - всё такое оборудование уже присутствует в месте назначения:
            #   + несколько позиций - часть оборудования - место где хоть какое-нибудь оборудование - часть такого оборудования уже присутствует в месте назначения
            #   + несколько позиций - часть оборудования - место где хоть какое-нибудь оборудование - такого оборудования там ещё нет
            # !!! если количество отправляемого оборудования равно количеству оборудования в БД (отправляем всё оборудование)
            if int(count_unit_send) == int(count_unit_db):
                # проверяем есть ли уже точно такое же оборудование (метод контроля, производитель, тип, название) в месте назначения
                # всё оборудование на локации
                all_unit_location = Unit.objects.filter(location=Location.objects.get(name=recip))
                # logging.error(f'all_unit_location {all_unit_location}')
                # !!! передаём всё оборудование на локацию, где есть хоть какое-нибудь оборудование
                if len(all_unit_location) != 0:
                    for i in all_unit_location:
                        # logging.error(f'Оборудование {i}')
                        no_unit_location = True
                        # !!! проверяем есть ли такое оборудование в месте назначения
                        # исключаем из обхода само перемещаемое оборудование
                        if not int(i.id) == int(index_id):
                            # logging.error(f'i.id {i.id}')
                            # logging.error(f'index_id {index_id}')
                            # если совпадают метод оборудования на локации и перемещаемого оборудования
                            if i.method == Unit.objects.get(id=int(index_id)).method:
                                # если совпадают производитель оборудования на локации и перемещаемого оборудования
                                if i.manufacturer == Unit.objects.get(id=int(index_id)).manufacturer:
                                    # если совпадают тип оборудования на локации и перемещаемого оборудования
                                    if i.type == Unit.objects.get(id=int(index_id)).type:
                                        # если совпадают название оборудования на локации и перемещаемого оборудования
                                        if i.equipment_name == Unit.objects.get(id=int(index_id)).equipment_name:
                                            # !!! передаваемое оборудование уже есть в месте назначения
                                            # увеличиваем количество на локации
                                            # logging.error('1')
                                            # try:
                                            #     logging.error(i.total)
                                            #     logging.error(Unit.objects.get(id=int(index_id)).total)
                                            # except Exception as e:
                                            #     logging.error(f'ошибка {e}')
                                            data_write[
                                                'total_write_coming'] = f'{int(i.total)} &#x2192 {int(i.total) + int(Unit.objects.get(id=int(index_id)).total)}'
                                            data_write[
                                                'total_write_expense'] = f'{int(Unit.objects.get(id=int(index_id)).total)} &#x2192 {int(Unit.objects.get(id=int(index_id)).total) - int(count_unit_send)}'
                                            i.total = int(i.total) + int(Unit.objects.get(id=int(index_id)).total)
                                            i.save()
                                            # logging.error('2')
                                            # предварительная запись откуда отправляется оборудование
                                            data_write['from_write'] = Unit.objects.get(id=int(index_id)).location
                                            # logging.error('3')
                                            data_write['location_write'] = recip
                                            # logging.error('4')
                                            # делаем запись истории в БД
                                            try:
                                                data_write['date_write'] = datetime.datetime.now().strftime('%Y-%m-%d')
                                                # logging.error('5')
                                                data_write['time_write'] = datetime.datetime.now().time().strftime('%H:%M:%S')
                                                # logging.error('6')
                                                data_write['crud_write_coming'] = f'получено(-а): {Unit.objects.get(id=int(index_id)).total} шт.'
                                                data_write['crud_write_expense'] = f'отправлено(-а): {Unit.objects.get(id=int(index_id)).total} шт.'
                                                # logging.error('7')
                                                data_write['who_write'] = request.user.username
                                                # Предварительная запись откуда отправляется оборудование - см. выше
                                                data_write['whom_write'] = recip_name
                                                data_write['to_write'] = recip
                                                data_write['method_write'] = Unit.objects.get(id=int(index_id)).method
                                                data_write['manufacturer_write'] = Unit.objects.get(id=int(index_id)).manufacturer
                                                data_write['type_write'] = Unit.objects.get(id=int(index_id)).type
                                                data_write['name_write'] = Unit.objects.get(id=int(index_id)).equipment_name
                                                data_write['serial_number_write'] = Unit.objects.get(id=int(index_id)).equipment_serial_number
                                                # Предварительная запись откуда &#x2192 куда - см. выше
                                                data_write['status_write'] = Unit.objects.get(id=int(index_id)).status
                                                data_write['notes_write'] = Unit.objects.get(id=int(index_id)).notes
                                                # logging.error('8')
                                                data_write['id_write'] = Unit.objects.get(id=int(index_id)).first_id
                                                # logging.error('9')
                                                # удаляем оборудование, которое отправляем
                                                # logging.error(f'Удаляемое оборудование {Unit.objects.get(id=int(index_id))}')
                                                Unit.objects.get(id=int(index_id)).delete()
                                                # logging.error('10')
                                                no_unit_location = False
                                                break
                                                # logging(f'1-1 {data_write}')
                                            except Exception as e:
                                                logging.error(f'1111 {e}')
                                                logging(f'1-2 {data_write}')
                    # !!! в месте назначения ещё нет такого оборудования
                    if no_unit_location:
                        # предварительная запись откуда отправляется оборудование
                        data_write['from_write'] = Unit.objects.get(id=int(index_id)).location
                        data_write['location_write'] = recip
                        # Предварительная запись было &#x2192 стало количество
                        data_write['total_write_coming'] = f'0 &#x2192 {count_unit_send}'
                        data_write[
                            'total_write_expense'] = f'{int(Unit.objects.get(id=int(index_id)).total)} &#x2192 {int(Unit.objects.get(id=int(index_id)).total) - int(count_unit_send)}'
                        # data_write['total_write'] = f'0 &#x2192 {int(i.total)}'
                        unit = Unit.objects.get(id=index_id)
                        location = Location.objects.get(name=recip)
                        unit.location = location
                        data_write['id_write'] = Unit.objects.get(id=int(index_id)).first_id
                        unit.save()
                        # делаем запись истории в БД
                        try:
                            data_write['date_write'] = datetime.datetime.now().strftime('%Y-%m-%d')
                            data_write['time_write'] = datetime.datetime.now().time().strftime('%H:%M:%S')
                            data_write['crud_write_coming'] = f'получено(-а): {Unit.objects.get(id=int(index_id)).total} шт.'
                            data_write['crud_write_expense'] = f'отправлено(-а): {Unit.objects.get(id=int(index_id)).total} шт.'
                            data_write['who_write'] = request.user.username
                            # Предварительная запись откуда отправляется оборудование - см. выше
                            data_write['whom_write'] = recip_name
                            data_write['to_write'] = recip
                            data_write['method_write'] = Unit.objects.get(id=int(index_id)).method
                            data_write['manufacturer_write'] = Unit.objects.get(id=int(index_id)).manufacturer
                            data_write['type_write'] = Unit.objects.get(id=int(index_id)).type
                            data_write['name_write'] = Unit.objects.get(id=int(index_id)).equipment_name
                            data_write['serial_number_write'] = Unit.objects.get(id=int(index_id)).equipment_serial_number
                            # Предварительная запись откуда &#x2192 куда - см. выше
                            data_write['status_write'] = Unit.objects.get(id=int(index_id)).status
                            data_write['notes_write'] = Unit.objects.get(id=int(index_id)).notes
                            # logging.error(f'2-1 {data_write}')
                        except Exception as e:
                            logging.error(f'2 {e}')
                            logging.error(f'2-2 {data_write}')
                # !!! передаём всё оборудование на локацию, где ещё нет никакого оборудования
                else:
                    # предварительная запись откуда отправляется оборудование
                    data_write['from_write'] = Unit.objects.get(id=int(index_id)).location
                    data_write['location_write'] = recip
                    unit = Unit.objects.get(id=index_id)
                    location = Location.objects.get(name=recip)
                    unit.location = location
                    unit.save()
                    # делаем запись истории в БД
                    try:
                        data_write['date_write'] = datetime.datetime.now().strftime('%Y-%m-%d')
                        data_write['time_write'] = datetime.datetime.now().time().strftime('%H:%M:%S')
                        data_write['crud_write_coming'] = f'получено(-а): {Unit.objects.get(id=int(index_id)).total} шт.'
                        data_write['crud_write_expense'] = f'отправлено(-а): {Unit.objects.get(id=int(index_id)).total} шт.'
                        data_write['who_write'] = request.user.username
                        # Предварительная запись откуда отправляется оборудование - см. выше
                        data_write['whom_write'] = recip_name
                        data_write['to_write'] = recip
                        data_write['method_write'] = Unit.objects.get(id=int(index_id)).method
                        data_write['manufacturer_write'] = Unit.objects.get(id=int(index_id)).manufacturer
                        data_write['type_write'] = Unit.objects.get(id=int(index_id)).type
                        data_write['name_write'] = Unit.objects.get(id=int(index_id)).equipment_name
                        data_write['serial_number_write'] = Unit.objects.get(id=int(index_id)).equipment_serial_number
                        data_write['total_write_coming'] = f'0 &#x2192 {int(Unit.objects.get(id=int(index_id)).total)}'
                        data_write[
                            'total_write_expense'] = f'{int(Unit.objects.get(id=int(index_id)).total)} &#x2192 {int(Unit.objects.get(id=int(index_id)).total) - int(count_unit_send)}'
                        # Предварительная запись откуда &#x2192 куда - см. выше
                        data_write['status_write'] = Unit.objects.get(id=int(index_id)).status
                        data_write['notes_write'] = Unit.objects.get(id=int(index_id)).notes
                        data_write['id_write'] = Unit.objects.get(id=int(index_id)).first_id
                        # logging.error(f'3-1 {data_write}')
                    except Exception as e:
                        logging.error(f'3 {e}')
                        logging.error(f'3-2 {data_write}')
            # !!! если количество отправляемого оборудования меньше чем количество оборудования в БД
            else:
                # проверяем есть ли уже точно такое же оборудование (метод контроля, производитель, тип, название) в месте назначения
                # всё оборудование на локации
                all_unit_location = Unit.objects.filter(location=Location.objects.get(name=recip))
                no_unit_location = True
                # !!! передаём часть оборудования на локацию, где есть хоть какое-нибудь оборудование
                if len(all_unit_location) != 0:
                    # logging.error(f'------------------------------------------------------------------------------------------')
                    # logging.error(f'Старт')
                    # logging.error(f'Всё оборудование на локации {all_unit_location}')
                    for i in all_unit_location:
                        # logging.error(f'i.id {i.id}')
                        # logging.error(f'index_id {index_id}')
                        # !!! проверяем есть ли такое оборудование в месте назначения
                        # исключаем из обхода само перемещаемое оборудование
                        if int(i.id) != int(index_id):
                            # logging.error(f'1')
                            # если совпадают метод оборудования на локации и перемещаемого оборудования
                            if i.method == Unit.objects.get(id=int(index_id)).method:
                                # logging.error(f'2')
                                # если совпадают производитель оборудования на локации и перемещаемого оборудования
                                if i.manufacturer == Unit.objects.get(id=int(index_id)).manufacturer:
                                    # logging.error(f'3')
                                    # если совпадают тип оборудования на локации и перемещаемого оборудования
                                    if i.type == Unit.objects.get(id=int(index_id)).type:
                                        # logging.error(f'4')
                                        # если совпадают название оборудования на локации и перемещаемого оборудования
                                        if i.equipment_name == Unit.objects.get(id=int(index_id)).equipment_name:
                                            # logging.error(f'Оборудование на локации: {i.equipment_name}')
                                            # logging.error(f'Количество оборудования на локации: {i.total}')
                                            # logging.error(f'Передаваемое оборудование: {Unit.objects.get(id=int(index_id)).equipment_name}')
                                            # logging.error(f'Количество передаваемого оборудования: {int(count_unit_send)}')
                                            # !!! передаваемое оборудование уже есть в месте назначения
                                            # Предварительная запись было &#x2192 стало количество
                                            data_write['total_write_coming'] = f'{int(i.total)} &#x2192 {int(i.total) + int(count_unit_send)}'
                                            data_write[
                                                'total_write_expense'] = f'{int(Unit.objects.get(id=int(index_id)).total)} &#x2192 {int(Unit.objects.get(id=int(index_id)).total) - int(count_unit_send)}'
                                            # увеличиваем количество на локации
                                            i.total = int(i.total) + int(count_unit_send)
                                            i.save()
                                            # logging.error(f'В итоге на локации: {i.total}')
                                            # делаем запись истории в БД
                                            try:
                                                data_write['date_write'] = datetime.datetime.now().strftime('%Y-%m-%d')
                                                # logging.error('part 5')
                                                data_write['time_write'] = datetime.datetime.now().time().strftime('%H:%M:%S')
                                                # logging.error('part 6')
                                                data_write['crud_write_coming'] = f'получено(-а): {int(count_unit_send)} шт.'
                                                data_write['crud_write_expense'] = f'отправлено(-а): {int(count_unit_send)} шт.'
                                                # logging.error('part 7')
                                                data_write['who_write'] = request.user.username
                                                data_write['from_write'] = Unit.objects.get(id=int(index_id)).location
                                                data_write['location_write'] = recip
                                                data_write['whom_write'] = recip_name
                                                data_write['to_write'] = recip
                                                data_write['method_write'] = Unit.objects.get(id=int(index_id)).method
                                                data_write['manufacturer_write'] = Unit.objects.get(id=int(index_id)).manufacturer
                                                data_write['type_write'] = Unit.objects.get(id=int(index_id)).type
                                                data_write['name_write'] = Unit.objects.get(id=int(index_id)).equipment_name
                                                data_write['serial_number_write'] = Unit.objects.get(id=int(index_id)).equipment_serial_number
                                                # Предварительная запись было &#x2192 стало количество - см. выше
                                                data_write['status_write'] = Unit.objects.get(id=int(index_id)).status
                                                data_write['notes_write'] = Unit.objects.get(id=int(index_id)).notes
                                                # logging.error('part 8')
                                                data_write['id_write'] = Unit.objects.get(id=int(index_id)).first_id
                                                # logging.error('part 9')
                                                no_unit_location = False
                                                # logging.error(f'10')
                                            except Exception as e:
                                                logging.error(e)
                                                logging.error(f'part 10 {data_write}')
                                            logging.error(f'11')
                                            unit_minus = Unit.objects.get(id=int(index_id))
                                            unit_minus.total = int(Unit.objects.get(id=int(index_id)).total) - int(count_unit_send)
                                            unit_minus.save()
                                            logging.error(f'12')
                                            logging.error(f'no_unit_location: {no_unit_location}')
                    # !!! в месте назначения ещё нет такого оборудования
                    if no_unit_location:
                        # копируем сведения о перемещаемом оборудовании
                        method = Unit.objects.get(id=index_id).method
                        manufacturer = Unit.objects.get(id=index_id).manufacturer
                        type = Unit.objects.get(id=index_id).type
                        equipment_name = Unit.objects.get(id=index_id).equipment_name
                        equipment_serial_number = Unit.objects.get(id=index_id).equipment_serial_number
                        location = Location.objects.get(name=recip)
                        # count_unit_send - количество оборудования
                        status = Unit.objects.get(id=index_id).status
                        notes = Unit.objects.get(id=index_id).notes
                        # создаём новое оборудование
                        unit = Unit()
                        unit.method = method
                        unit.manufacturer = manufacturer
                        unit.type = type
                        unit.equipment_name = equipment_name
                        unit.equipment_serial_number = equipment_serial_number
                        unit.total = count_unit_send
                        unit.location = location
                        unit.status = status
                        unit.notes = notes
                        unit.first_id = Unit.objects.get(id=int(index_id)).first_id
                        unit.save()
                        # Предварительная запись было &#x2192 стало количество
                        data_write['total_write_coming'] = f'0 &#x2192 {int(count_unit_send)}'
                        data_write[
                            'total_write_expense'] = f'{int(Unit.objects.get(id=int(index_id)).total)} &#x2192 {int(Unit.objects.get(id=int(index_id)).total) - int(count_unit_send)}'
                        # уменьшаем количество отправленного оборудования в локации откуда отправляется оборудование
                        unit = Unit.objects.get(id=index_id)
                        unit.total = int(count_unit_db) - int(count_unit_send)
                        data_write['id_write'] = Unit.objects.get(id=int(index_id)).first_id
                        unit.save()
                        # делаем запись истории в БД
                        try:
                            data_write['date_write'] = datetime.datetime.now().strftime('%Y-%m-%d')
                            # logging.error('part 15')
                            data_write['time_write'] = datetime.datetime.now().time().strftime('%H:%M:%S')
                            # logging.error('part 16')
                            data_write['crud_write_coming'] = f'получено(-а): {int(count_unit_send)} шт.'
                            data_write['crud_write_expense'] = f'отправлено(-а): {int(count_unit_send)} шт.'
                            # logging.error('part 17')
                            data_write['who_write'] = request.user.username
                            data_write['from_write'] = Unit.objects.get(id=int(index_id)).location
                            data_write['location_write'] = recip
                            data_write['whom_write'] = recip_name
                            data_write['to_write'] = recip
                            data_write['method_write'] = Unit.objects.get(id=int(index_id)).method
                            data_write['manufacturer_write'] = Unit.objects.get(id=int(index_id)).manufacturer
                            data_write['type_write'] = Unit.objects.get(id=int(index_id)).type
                            data_write['name_write'] = Unit.objects.get(id=int(index_id)).equipment_name
                            data_write['serial_number_write'] = Unit.objects.get(id=int(index_id)).equipment_serial_number
                            # Предварительная запись было &#x2192 стало количество - см. выше
                            data_write['status_write'] = Unit.objects.get(id=int(index_id)).status
                            data_write['notes_write'] = Unit.objects.get(id=int(index_id)).notes
                            # logging.error('part 18')
                            # logging.error('part 19')
                        except Exception as e:
                            logging.error(e)
                            logging.error(f'part 110 {data_write}')
                # !!! передаём часть оборудования на локацию, где ещё нет никакого оборудования
                else:
                    # копируем сведения о перемещаемом оборудовании
                    method = Unit.objects.get(id=index_id).method
                    manufacturer = Unit.objects.get(id=index_id).manufacturer
                    type = Unit.objects.get(id=index_id).type
                    equipment_name = Unit.objects.get(id=index_id).equipment_name
                    equipment_serial_number = Unit.objects.get(id=index_id).equipment_serial_number
                    location = Location.objects.get(name=recip)
                    # count_unit_send - количество оборудования
                    status = Unit.objects.get(id=index_id).status
                    notes = Unit.objects.get(id=index_id).notes
                    # создаём новое оборудование
                    unit = Unit()
                    unit.method = method
                    unit.manufacturer = manufacturer
                    unit.type = type
                    unit.equipment_name = equipment_name
                    unit.equipment_serial_number = equipment_serial_number
                    unit.total = count_unit_send
                    unit.location = location
                    unit.status = status
                    unit.notes = notes
                    unit.first_id = Unit.objects.get(id=index_id).first_id
                    unit.save()
                    # Предварительная запись было &#x2192 стало количество
                    data_write['total_write_coming'] = f'0 &#x2192 {int(count_unit_send)}'
                    data_write[
                        'total_write_expense'] = f'{int(Unit.objects.get(id=int(index_id)).total)} &#x2192 {int(Unit.objects.get(id=int(index_id)).total) - int(count_unit_send)}'
                    data_write['id_write'] = Unit.objects.get(id=int(index_id)).first_id
                    # уменьшаем количество отправленного оборудования в локации откуда отправляется оборудование
                    unit = Unit.objects.get(id=index_id)
                    unit.total = int(count_unit_db) - int(count_unit_send)
                    unit.save()
                    # делаем запись истории в БД
                    try:
                        data_write['date_write'] = datetime.datetime.now().strftime('%Y-%m-%d')
                        # logging.error('part 15')
                        data_write['time_write'] = datetime.datetime.now().time().strftime('%H:%M:%S')
                        # logging.error('part 16')
                        data_write['crud_write_coming'] = f'получено(-а): {int(count_unit_send)} шт.'
                        data_write['crud_write_expense'] = f'отправлено(-а): {int(count_unit_send)} шт.'
                        # logging.error('part 17')
                        data_write['who_write'] = request.user.username
                        data_write['from_write'] = Unit.objects.get(id=int(index_id)).location
                        data_write['location_write'] = recip
                        data_write['whom_write'] = recip_name
                        data_write['to_write'] = recip
                        data_write['method_write'] = Unit.objects.get(id=int(index_id)).method
                        data_write['manufacturer_write'] = Unit.objects.get(id=int(index_id)).manufacturer
                        data_write['type_write'] = Unit.objects.get(id=int(index_id)).type
                        data_write['name_write'] = Unit.objects.get(id=int(index_id)).equipment_name
                        data_write['serial_number_write'] = Unit.objects.get(id=int(index_id)).equipment_serial_number
                        # Предварительная запись было &#x2192 стало количество - см. выше
                        data_write['status_write'] = Unit.objects.get(id=int(index_id)).status
                        data_write['notes_write'] = Unit.objects.get(id=int(index_id)).notes
                        # logging.error('part 18')
                        # logging.error('part 19')
                    except Exception as e:
                        logging.error(e)
                        logging.error(f'part 110 {data_write}')
            try:
                logging.error(f'Финиш')
                write_history_coming(request, data_write)
                write_history_expense(request, data_write)
            except Exception as e:
                logger.error(f'2 {e}')
        return HttpResponse()
    else:
        return HttpResponse("Метод запроса должен быть POST.", status=405)


@permission_required('catalog.can_delete_task', raise_exception=True)
def unit_delete(request, id):
    unit = Unit.objects.get(id=id)
    # словарь со значениями для формирования истории в БД
    data_write = {'date_write': datetime.datetime.now().strftime('%Y-%m-%d'),
                  'time_write': datetime.datetime.now().time().strftime('%H:%M:%S'),
                  'crud_write': f'удалено из БД',
                  'from_write': '-',
                  'who_write': request.user.username,
                  'whom_write': '-',
                  'to_write': '-',
                  'method_write': unit.method,
                  'manufacturer_write': unit.manufacturer,
                  'type_write': unit.type,
                  'name_write': unit.equipment_name,
                  'serial_number_write': unit.equipment_serial_number,
                  'total_write': unit.total,
                  'location_write': unit.location,
                  # 'location_write':  Location.objects.get(id=request.POST.get("location")),
                  'status_write': unit.status,
                  # 'status_write': Status.objects.get(id=request.POST.get("status")),
                  'notes_write': unit.notes,
                  'id_write':  unit.first_id
                  }
    unit.delete()
    write_history(request, data_write)
    return HttpResponseRedirect('/home/#delete-tab-pane')
    # return HttpResponseRedirect('/home/')


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


@permission_required('catalog.can_send_task', raise_exception=True)
def forms_send(request):
    queryset = Unit.objects.all()
    unit = {}
    names = Location.objects.values_list('name', flat=True)
    loc = list(names)
    for index, i in enumerate(queryset):
        unit[index] = {'id': i.id, 'name': i.equipment_name, 'serial': i.equipment_serial_number, 'total': i.total, 'type': i.type,
                       'manufacturer': i.manufacturer, 'location': loc, 'loc': i.location.name}
    data = json.dumps(unit)
    return render(request, "catalog/forms_send.html", {'data': data})


@csrf_exempt  # Отключение CSRF для упрощения (не используйте в production)
def send_excel(request):
    if request.method == 'POST':
        try:
            # список id оборудования для отправки
            id_list = []
            raw_body = request.body.decode('utf-8')
            data = json.loads(raw_body)
            # Извлекаем данные
            table_data = data.get('data', [])
            # создаём файл Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Form Data"

            # вписать лист в одну страницу
            ws.page_setup.fitToPage = True
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            # Масштабирование по ширине (1 страница в ширину)
            ws.page_setup.fitToWidth = 1
            # Не ограничивать количество страниц по высоте
            ws.page_setup.fitToHeight = 0
            # чёрный стиль для границы
            thin = Side(border_style="thin", color="000000")
            medium = Side(border_style="medium", color="000000")
            # вставка изображение и адрес в зависимости от выбранного департамента
            if table_data[-1]['departament'] == 'YKR':
                img = Image(os.path.join(settings.BASE_DIR, 'catalog', 'static', 'images', 'Rutledge.png'))
                # img = Image(os.path.join(settings.BASE_DIR, 'staticfiles', 'images', 'Rutledge.png')) продакшн
                ws.add_image(img, 'C3')
                ws['E3'] = '      Yeskert Kyzmet Rutledge LLP,'
                ws['E3'].font = Font(name='Times New Roman', size=9)
                ws['E4'] = '      Atyrau, Republic of Kazakhstan,'
                ws['E4'].font = Font(name='Times New Roman', size=9)
                ws['E5'] = '      NDT department'
                ws['E5'].font = Font(name='Times New Roman', size=9)
            if table_data[-1]['departament'] == 'Arise':
                img = Image(os.path.join(settings.BASE_DIR, 'catalog', 'static', 'images', 'Arise.png'))
                ws.add_image(img, 'C3')
                ws['E3'] = '      Arise Kazakhstan LLP,'
                ws['E3'].font = Font(name='Times New Roman', size=9)
                ws['E4'] = '      Atyrau, Republic of Kazakhstan,'
                ws['E4'].font = Font(name='Times New Roman', size=9)
                ws['E5'] = '      NDT department'
                ws['E5'].font = Font(name='Times New Roman', size=9)
            # шапка листа с форматированием
            ws['F3'] = 'Date:'
            ws['F3'].alignment = Alignment(horizontal='right')
            ws['F3'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['F4'] = 'Time:'
            ws['F4'].alignment = Alignment(horizontal='right')
            ws['F4'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['G3'] = f'{datetime.datetime.now():%A, %d %B %Y}'
            ws['G3'].font = Font(name='Times New Roman', size=9, underline='single')
            ws['G4'] = f'{datetime.datetime.now().time():%H:%M}'
            ws['G4'].font = Font(name='Times New Roman', size=9, underline='single')
            ws['E8'] = 'EQUIPMENT TRANSFER AGREEMENT'
            ws['E8'].alignment = Alignment(horizontal='center')
            ws['E8'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['E9'] = 'Акт приема-передачи оборудования'
            ws['E9'].alignment = Alignment(horizontal='center')
            ws['E9'].font = Font(name='Times New Roman', size=11)
            # объединяем две ячейки для "№"
            ws.merge_cells(start_row=12, start_column=2, end_row=13, end_column=2)
            # заголовки таблицы на английском с форматированием
            ws['B12'] = '№'
            ws['B12'].alignment = Alignment(horizontal='center', vertical='center')
            ws['B12'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['C12'] = 'Type'
            ws['C12'].alignment = Alignment(horizontal='center')
            ws['C12'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['D12'] = 'Manufacture'
            ws['D12'].alignment = Alignment(horizontal='center')
            ws['D12'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['E12'] = 'Name'
            ws['E12'].alignment = Alignment(horizontal='center')
            ws['E12'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['F12'] = 'Serial Number'
            ws['F12'].alignment = Alignment(horizontal='center')
            ws['F12'].font = Font(name='Times New Roman', size=11, bold=True)
            ws['G12'] = 'Quantity'
            ws['G12'].alignment = Alignment(horizontal='center')
            ws['G12'].font = Font(name='Times New Roman', size=11, bold=True)
            # заголовки таблицы на русском с форматированием
            ws['B13'].alignment = Alignment(horizontal='center')
            ws['C13'] = 'Тип'
            ws['C13'].alignment = Alignment(horizontal='center')
            ws['C13'].font = Font(name='Times New Roman', size=11)
            ws['D13'] = 'Производитель'
            ws['D13'].alignment = Alignment(horizontal='center')
            ws['D13'].font = Font(name='Times New Roman', size=11)
            ws['E13'] = 'Наименование'
            ws['E13'].alignment = Alignment(horizontal='center')
            ws['E13'].font = Font(name='Times New Roman', size=11)
            ws['F13'] = 'Серийный номер'
            ws['F13'].alignment = Alignment(horizontal='center')
            ws['F13'].font = Font(name='Times New Roman', size=11)
            ws['G13'] = 'Кол-во'
            ws['G13'].alignment = Alignment(horizontal='center')
            ws['G13'].font = Font(name='Times New Roman', size=11)
            # высота первой строки
            ws.row_dimensions[1].height = 5.40
            # ширина столбцов
            ws.column_dimensions['A'].width = 2.33
            ws.column_dimensions['B'].width = 3.67
            ws.column_dimensions['C'].width = 14.67
            ws.column_dimensions['D'].width = 14.89
            ws.column_dimensions['E'].width = 48.78
            ws.column_dimensions['F'].width = 16.89
            ws.column_dimensions['G'].width = 10.22
            ws.column_dimensions['H'].width = 11.78
            # последняя строка основных данных таблицы
            last_index = 0
            # общее количество оборудования для отправки
            last_count_send_equipment = 0
            # основные данные таблицы
            for index, item in enumerate(table_data):
                if not len(item) == 3:
                    ws.row_dimensions[14 + index].height = 24.60
                    ws[f'B{14 + index}'] = int(item['index'])
                    ws[f'B{14 + index}'].number_format = numbers.FORMAT_NUMBER
                    ws[f'B{14 + index}'].border = Border(top=thin, bottom=thin, left=medium, right=thin)
                    ws[f'C{14 + index}'] = item['type']
                    ws[f'C{14 + index}'].border = Border(top=thin, bottom=thin, left=thin, right=thin)
                    ws[f'D{14 + index}'] = item['manufacturer']
                    ws[f'D{14 + index}'].border = Border(top=thin, bottom=thin, left=thin, right=thin)
                    ws[f'E{14 + index}'] = item['name']
                    ws[f'E{14 + index}'].border = Border(top=thin, bottom=thin, left=thin, right=thin)
                    ws[f'F{14 + index}'] = item['serial']
                    ws[f'F{14 + index}'].border = Border(top=thin, bottom=thin, left=thin, right=thin)
                    ws[f'G{14 + index}'] = int(item['count'])
                    ws[f'G{14 + index}'].border = Border(top=thin, bottom=thin, left=thin, right=medium)
                    last_count_send_equipment += int(item['count'])
                    # форматирование основных данных
                    ws[f'B{14 + index}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'B{14 + index}'].font = Font(name='Bahnschrift SemiLight', size=10)
                    ws[f'C{14 + index}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws[f'C{14 + index}'].font = Font(name='Bahnschrift SemiLight', size=10)
                    ws[f'D{14 + index}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws[f'D{14 + index}'].font = Font(name='Bahnschrift SemiLight', size=10)
                    ws[f'E{14 + index}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws[f'E{14 + index}'].font = Font(name='Bahnschrift SemiLight', size=10)
                    ws[f'F{14 + index}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'F{14 + index}'].font = Font(name='Bahnschrift SemiLightn', size=10)
                    ws[f'G{14 + index}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws[f'G{14 + index}'].font = Font(name='Bahnschrift SemiLight', size=10)
                    id_list.append(item['id'])
                last_index = 14 + index
            # граница шапки таблицы
            ws[f'B12'].border = Border(top=medium, bottom=medium, left=medium)
            ws[f'B13'].border = Border(top=medium, bottom=medium, left=medium)
            ws[f'C12'].border = Border(top=medium, left=thin, right=thin)
            ws[f'C13'].border = Border(bottom=medium, left=thin, right=thin)
            ws[f'D12'].border = Border(top=medium, left=thin, right=thin)
            ws[f'D13'].border = Border(bottom=medium, left=thin, right=thin)
            ws[f'E12'].border = Border(top=medium, left=thin, right=thin)
            ws[f'E13'].border = Border(bottom=medium, left=thin, right=thin)
            ws[f'F12'].border = Border(top=medium, left=thin, right=thin)
            ws[f'F13'].border = Border(bottom=medium, left=thin, right=thin)
            ws[f'G12'].border = Border(top=medium, left=thin, right=medium)
            ws[f'G13'].border = Border(bottom=medium, left=thin, right=medium)
            # граница футера таблицы
            ws[f'B{last_index}'].border = Border(top=medium)
            ws[f'C{last_index}'].border = Border(top=medium)
            ws[f'D{last_index}'].border = Border(top=medium)
            ws[f'E{last_index}'].border = Border(top=medium)
            ws[f'F{last_index}'].border = Border(top=medium)
            ws[f'G{last_index}'].border = Border(top=medium)

            # футер листа с форматированием
            # высота первой строки
            ws.row_dimensions[last_index].height = 27
            ws[f'G{int(last_index)}'] = last_count_send_equipment
            ws[f'G{int(last_index)}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'G{int(last_index)}'].font = Font(name='Times New Roman', size=11)
            ws[f'G{int(last_index)}'].border = Border(top=medium, bottom=medium, left=medium, right=medium)

            ws.merge_cells(start_row=int(last_index) + 2, start_column=2, end_row=int(last_index) + 2, end_column=3)
            ws[f'B{int(last_index) + 2}'] = 'Handed out:'
            ws[f'B{int(last_index) + 2}'].font = Font(name='Times New Roman', size=11, bold=True)

            ws.merge_cells(start_row=int(last_index) + 3, start_column=2, end_row=int(last_index) + 3, end_column=3)
            ws[f'B{int(last_index) + 3}'] = 'Сдал'
            ws[f'B{int(last_index) + 3}'].font = Font(name='Times New Roman', size=11)

            ws.merge_cells(start_row=int(last_index) + 5, start_column=2, end_row=int(last_index) + 5, end_column=3)
            ws[f'B{int(last_index) + 5}'] = 'Sign (подпись)'
            ws[f'B{int(last_index) + 5}'].alignment = Alignment(horizontal='center')
            ws[f'B{int(last_index) + 5}'].font = Font(name='Times New Roman', size=8)
            # thin = Side(border_style="thin", color="000000")
            ws[f'B{int(last_index) + 5}'].border = Border(top=thin)
            ws[f'C{int(last_index) + 5}'].border = Border(top=thin)

            ws.merge_cells(start_row=int(last_index) + 5, start_column=4, end_row=int(last_index) + 5, end_column=5)
            ws[f'D{int(last_index) + 5}'] = 'Full name (расшифровка подписи)'
            ws[f'D{int(last_index) + 5}'].alignment = Alignment(horizontal='center')
            ws[f'D{int(last_index) + 5}'].font = Font(name='Times New Roman', size=8)
            ws[f'D{int(last_index) + 5}'].border = Border(top=thin)
            ws[f'E{int(last_index) + 5}'].border = Border(top=thin)

            ws.merge_cells(start_row=int(last_index) + 7, start_column=2, end_row=int(last_index) + 7, end_column=3)
            ws[f'B{int(last_index) + 7}'] = 'Received: '
            ws[f'B{int(last_index) + 7}'].font = Font(name='Times New Roman', size=11, bold=True)

            ws.merge_cells(start_row=int(last_index) + 8, start_column=2, end_row=int(last_index) + 8, end_column=3)
            ws[f'B{int(last_index) + 8}'] = 'Принял'
            ws[f'B{int(last_index) + 8}'].font = Font(name='Times New Roman', size=11)

            ws.merge_cells(start_row=int(last_index) + 10, start_column=2, end_row=int(last_index) + 10, end_column=3)
            ws[f'B{int(last_index) + 10}'] = 'Sign (подпись)'
            ws[f'B{int(last_index) + 10}'].alignment = Alignment(horizontal='center')
            ws[f'B{int(last_index) + 10}'].font = Font(name='Times New Roman', size=8)

            ws[f'B{int(last_index) + 10}'].border = Border(top=thin)
            ws[f'C{int(last_index) + 10}'].border = Border(top=thin)

            ws.merge_cells(start_row=int(last_index) + 10, start_column=4, end_row=int(last_index) + 10, end_column=5)
            ws[f'D{int(last_index) + 10}'] = 'Full name (расшифровка подписи)'
            ws[f'D{int(last_index) + 10}'].alignment = Alignment(horizontal='center')
            ws[f'D{int(last_index) + 10}'].font = Font(name='Times New Roman', size=8)
            ws[f'D{int(last_index) + 10}'].border = Border(top=thin)
            ws[f'E{int(last_index) + 10}'].border = Border(top=thin)

            # Сохранение файла в буфер
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            # Проверяем какое количество оборудования перемещается и изменяем значение "Местоположение"
            check_count_send_equipment(request, id_list, table_data)

            # Возврат файла как ответ
            response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=output.xlsx'
            return response
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'Invalid request method'}, status=400)


# делаем запись в историю БД об отправки (приход) оборудования
def write_history_coming(request, data_write: dict):
    # if request.method == 'POST':
    crud_history = WriteHistory()
    try:
        # logging.error(f'4-1 {data_write}')
        crud_history.date_write = data_write['date_write']
        crud_history.time_write = data_write['time_write']
        crud_history.crud_write = data_write['crud_write_coming']
        crud_history.who_write = data_write['who_write']
        crud_history.from_write = data_write['from_write']
        crud_history.whom_write = data_write['whom_write']
        crud_history.to_write = data_write['to_write']
        crud_history.method_write = data_write['method_write']
        crud_history.manufacturer_write = data_write['manufacturer_write']
        crud_history.type_write = data_write['type_write']
        crud_history.name_write = data_write['name_write']
        crud_history.serial_number_write = data_write['serial_number_write']
        crud_history.total_write = data_write['total_write_coming']
        crud_history.location_write = data_write['location_write']
        crud_history.status_write = data_write['status_write']
        crud_history.notes_write = data_write['notes_write']
        crud_history.id_write = data_write['id_write']
    except Exception as e:
        logging.error(f'4 {e}')
        logging.error(f'4-2 {data_write}')
    crud_history.save()


# делаем запись в историю БД об отправки (расход) оборудования
def write_history_expense(request, data_write: dict):
    # if request.method == 'POST':
    crud_history = WriteHistory()
    try:
        # logging.error(f'4-1 {data_write}')
        crud_history.date_write = data_write['date_write']
        crud_history.time_write = data_write['time_write']
        crud_history.crud_write = data_write['crud_write_expense']
        crud_history.who_write = data_write['who_write']
        crud_history.from_write = data_write['from_write']
        crud_history.whom_write = data_write['whom_write']
        crud_history.to_write = data_write['to_write']
        crud_history.method_write = data_write['method_write']
        crud_history.manufacturer_write = data_write['manufacturer_write']
        crud_history.type_write = data_write['type_write']
        crud_history.name_write = data_write['name_write']
        crud_history.serial_number_write = data_write['serial_number_write']
        crud_history.total_write = data_write['total_write_expense']
        crud_history.location_write = data_write['from_write']
        crud_history.status_write = data_write['status_write']
        crud_history.notes_write = data_write['notes_write']
        crud_history.id_write = data_write['id_write']
    except Exception as e:
        logging.error(f'4 {e}')
        logging.error(f'4-2 {data_write}')
    crud_history.save()


# делаем запись в историю БД о создании оборудования
def write_history(request, data_write: dict):
    crud_history = WriteHistory()
    try:
        crud_history.date_write = data_write['date_write']
        crud_history.time_write = data_write['time_write']
        crud_history.crud_write = data_write['crud_write']
        crud_history.who_write = data_write['who_write']
        crud_history.from_write = data_write['from_write']
        crud_history.whom_write = data_write['whom_write']
        crud_history.to_write = data_write['to_write']
        crud_history.method_write = data_write['method_write']
        crud_history.manufacturer_write = data_write['manufacturer_write']
        crud_history.type_write = data_write['type_write']
        crud_history.name_write = data_write['name_write']
        crud_history.serial_number_write = data_write['serial_number_write']
        crud_history.total_write = data_write['total_write']
        crud_history.location_write = data_write['location_write']
        crud_history.status_write = data_write['status_write']
        crud_history.notes_write = data_write['notes_write']
        crud_history.id_write = data_write['id_write']
    except Exception as e:
        logging.error(f'4 {e}')
        logging.error(f'4-2 {data_write}')
    crud_history.save()


def get_rows_as_dicts(query, params=None):
    """
    Выполняет SQL-запрос и возвращает результат в виде списка словарей.
    """
    with connection.cursor() as cursor:
        cursor.execute(query, params or [])
        columns = [col[0] for col in cursor.description]  # Получаем имена столбцов
        rows = cursor.fetchall()
        # logging.error(f'0 rows {rows}')
    # Преобразуем строки в словари
    result = [dict(zip(columns, row)) for row in rows]
    # logging.error(f'1 result {result}')
    return result


def unit_history(request, first_id):
    if not first_id:
        return render(request, 'catalog/unit_history.html', {
            "error": "first_id is required",  # Сообщение об ошибке, если first_id не передан
        })

    # SQL-запрос для выборки строк по столбцу id_write
    query = """
        SELECT date_write, time_write, from_write, who_write, crud_write, to_write, whom_write, method_write, name_write, 
               serial_number_write, status_write, notes_write, location_write, total_write, type_write, manufacturer_write
        FROM catalog_writehistory
        WHERE id_write = %s
        ORDER BY date_write, time_write DESC
    """

    # Получаем данные из таблицы в виде списка словарей
    row_list = get_rows_as_dicts(query, [first_id])
    # logging.error(f'2 first_id {first_id}')
    # logging.error(f'3 row_list {row_list}')
    # Передаем данные в шаблон
    return render(request, 'catalog/unit_history.html', {"row_list": row_list})
